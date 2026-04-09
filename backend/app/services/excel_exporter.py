from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.mtr_card import MtrCard


def export_cards_to_excel(cards: list[MtrCard], char_names: list[str]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Результаты"

    headers = [
        "GUID",
        "Наименование",
        "ИНН изготовителя",
        "Артикул",
        "Цена EXW",
        "Дата начала цены",
        "Дата конца цены",
        "Описание",
        *char_names,
    ]
    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for card in cards:
        characteristics_map = {item.char_name: item for item in card.characteristics}
        row = [
            card.guid,
            card.name,
            card.manufacturer_inn,
            card.article,
            card.price,
            card.price_date_start,
            card.price_date_end,
            card.description,
        ]
        for char_name in char_names:
            characteristic = characteristics_map.get(char_name)
            row.append(_format_characteristic_value(characteristic))
        sheet.append(row)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float, Decimal)):
                cell.alignment = Alignment(horizontal="right")

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _format_characteristic_value(characteristic: object) -> object:
    if characteristic is None:
        return None
    if getattr(characteristic, "char_type", None) == "range" or getattr(characteristic, "range_min", None) is not None or getattr(characteristic, "range_max", None) is not None:
        range_min = _decimal_to_string(getattr(characteristic, "range_min", None))
        range_max = _decimal_to_string(getattr(characteristic, "range_max", None))
        if range_min is None and range_max is None:
            return None
        return f"{range_min or ''}/{range_max or ''}"
    if getattr(characteristic, "value_numeric", None) is not None:
        return getattr(characteristic, "value_numeric")
    return getattr(characteristic, "value_text", None)


def _decimal_to_string(value: Decimal | None) -> str | None:
    if value is None:
        return None
    normalized = value.normalize()
    text = format(normalized, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text
